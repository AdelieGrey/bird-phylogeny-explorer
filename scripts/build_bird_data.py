import csv
import json
import re
import sys
import urllib.request
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / ".local_py"))
import xlrd

RAW_DIR = ROOT / "data" / "raw"
APP_DIR = ROOT / "app"
CLEMENTS_PATH = RAW_DIR / "eBird-Clements_v2025-integrated-checklist-October-2025.csv"
AVILIST_PATH = RAW_DIR / "AviList-v2025b-10Jun2026-extended.xlsx"
CHINESE_PATH = RAW_DIR / "birdMapV2.js"
CHINA_CHECKLIST_PATH = RAW_DIR / "中国观鸟年报-中国鸟类名录_v12.0.xls"


ORDER_ZH = {
    "Accipitriformes": "鹰形目",
    "Aegotheliformes": "裸鼻鸱目",
    "Anseriformes": "雁形目",
    "Apodiformes": "雨燕目",
    "Apterygiformes": "几维目",
    "Bucerotiformes": "犀鸟目",
    "Caprimulgiformes": "夜鹰目",
    "Cariamiformes": "叫鹤目",
    "Casuariiformes": "鹤鸵目",
    "Cathartiformes": "美洲鹫目",
    "Charadriiformes": "鸻形目",
    "Ciconiiformes": "鹳形目",
    "Coliiformes": "鼠鸟目",
    "Columbiformes": "鸽形目",
    "Coraciiformes": "佛法僧目",
    "Cuculiformes": "鹃形目",
    "Eurypygiformes": "日鳽目",
    "Falconiformes": "隼形目",
    "Galliformes": "鸡形目",
    "Galbuliformes": "鴷形目",
    "Gaviiformes": "潜鸟目",
    "Gruiformes": "鹤形目",
    "Leptosomiformes": "鹃三宝鸟目",
    "Mesitornithiformes": "拟鹑目",
    "Musophagiformes": "蕉鹃目",
    "Nyctibiiformes": "林鸱目",
    "Opisthocomiformes": "麝雉目",
    "Otidiformes": "鸨形目",
    "Passeriformes": "雀形目",
    "Pelecaniformes": "鹈形目",
    "Phaethontiformes": "鹲形目",
    "Phoenicopteriformes": "红鹳目",
    "Piciformes": "䴕形目",
    "Podargiformes": "蟆口鸱目",
    "Podicipediformes": "䴙䴘目",
    "Procellariiformes": "鹱形目",
    "Psittaciformes": "鹦形目",
    "Pterocliformes": "沙鸡目",
    "Rheiformes": "美洲鸵目",
    "Sphenisciformes": "企鹅目",
    "Steatornithiformes": "油鸱目",
    "Strigiformes": "鸮形目",
    "Struthioniformes": "鸵鸟目",
    "Suliformes": "鲣鸟目",
    "Tinamiformes": "䳍形目",
    "Trogoniformes": "咬鹃目",
}


ORDER_PARENT = {
    "Struthioniformes": "palaeognathae",
    "Rheiformes": "palaeognathae",
    "Casuariiformes": "palaeognathae",
    "Apterygiformes": "palaeognathae",
    "Tinamiformes": "palaeognathae",
    "Galliformes": "galloanserae",
    "Anseriformes": "galloanserae",
    "Steatornithiformes": "strisores",
    "Nyctibiiformes": "strisores",
    "Podargiformes": "strisores",
    "Aegotheliformes": "strisores",
    "Caprimulgiformes": "strisores",
    "Apodiformes": "strisores",
    "Gaviiformes": "aequornithes",
    "Sphenisciformes": "aequornithes",
    "Procellariiformes": "aequornithes",
    "Ciconiiformes": "aequornithes",
    "Suliformes": "aequornithes",
    "Pelecaniformes": "aequornithes",
    "Phaethontiformes": "phaethontimorphae",
    "Phoenicopteriformes": "mirandornithes",
    "Podicipediformes": "mirandornithes",
    "Columbiformes": "columbaves",
    "Pterocliformes": "columbaves",
    "Mesitornithiformes": "columbaves",
    "Cuculiformes": "otidimorphae",
    "Musophagiformes": "otidimorphae",
    "Otidiformes": "otidimorphae",
    "Gruiformes": "cursorimorphae",
    "Charadriiformes": "cursorimorphae",
    "Eurypygiformes": "phaethontimorphae",
    "Opisthocomiformes": "gruae",
    "Leptosomiformes": "afroaves",
    "Coliiformes": "afroaves",
    "Trogoniformes": "afroaves",
    "Bucerotiformes": "afroaves",
    "Coraciiformes": "afroaves",
    "Galbuliformes": "afroaves",
    "Piciformes": "afroaves",
    "Cathartiformes": "afroaves",
    "Accipitriformes": "afroaves",
    "Strigiformes": "afroaves",
    "Cariamiformes": "australaves",
    "Falconiformes": "australaves",
    "Psittaciformes": "australaves",
    "Passeriformes": "australaves",
}

CLADE_SORT = {
    "aves": 0,
    "neornithes": 5,
    "palaeognathae": 10,
    "neognathae": 20,
    "galloanserae": 30,
    "neoaves": 40,
    "strisores": 100,
    "columbaves": 120,
    "otidimorphae": 130,
    "mirandornithes": 140,
    "elementaves": 150,
    "gruae": 151,
    "cursorimorphae": 152,
    "strisores_phaethoquornithes": 160,
    "phaethoquornithes": 161,
    "phaethontimorphae": 162,
    "aequornithes": 170,
    "telluraves": 200,
    "afroaves": 210,
    "australaves": 220,
}


CLADES = [
    {
        "id": "aves",
        "rank": "class",
        "scientificName": "Aves",
        "englishName": "Birds",
        "chineseName": "鸟纲",
        "parentId": None,
        "summary": "现代鸟类总节点。本原型从这里展开到传统目、科和精选物种。",
        "traits": ["羽毛、喙、产卵、前肢演化为翅", "飞行能力多次丢失或极端特化"],
    },
    {
        "id": "neornithes",
        "rank": "subclass",
        "scientificName": "Neornithes",
        "englishName": "Modern birds",
        "chineseName": "今鸟亚纲",
        "parentId": "aves",
        "summary": "现生鸟类及其近缘灭绝成员所在的冠群，向下分为古颚类与今颚类。",
        "traits": ["现代鸟类主要谱系的共同起点", "包含古颚类与今颚类两大分支"],
    },
    {
        "id": "palaeognathae",
        "rank": "clade",
        "scientificName": "Palaeognathae",
        "englishName": "Palaeognaths",
        "chineseName": "古颚类",
        "parentId": "neornithes",
        "summary": "包括鸵鸟、鸸鹋、鹤鸵、几维、䳍等，是现代鸟类早期分化出的主要分支之一。",
        "traits": ["多数大型陆栖或飞行能力弱", "包含不会飞的平胸鸟和会飞的䳍类"],
    },
    {
        "id": "neognathae",
        "rank": "clade",
        "scientificName": "Neognathae",
        "englishName": "Neognaths",
        "chineseName": "今颚类",
        "parentId": "neornithes",
        "summary": "除古颚类之外的现代鸟类大分支，包含鸡雁小纲与新鸟类。",
        "traits": ["现代鸟类绝大多数多样性所在", "生态位跨度极大"],
    },
    {
        "id": "galloanserae",
        "rank": "clade",
        "scientificName": "Galloanserae",
        "englishName": "Fowl and waterfowl",
        "chineseName": "鸡雁小纲",
        "parentId": "neognathae",
        "summary": "鸡形目与雁形目的共同分支，是今颚类中很早分出的一个大类群。",
        "traits": ["许多种类早成雏明显", "包含重要的地栖、涉水与水生鸟类"],
    },
    {
        "id": "neoaves",
        "rank": "clade",
        "scientificName": "Neoaves",
        "englishName": "Neoaves",
        "chineseName": "新鸟类",
        "parentId": "neognathae",
        "summary": "现代鸟类最大辐射分支之一，包含除古颚类和鸡雁小纲之外的大多数现生鸟类。",
        "traits": ["包含绝大多数现代鸟类目", "多个深部分支在短时间内快速辐射"],
    },
    {
        "id": "strisores",
        "rank": "clade",
        "scientificName": "Strisores",
        "englishName": "Nightbirds, swifts, and hummingbirds",
        "chineseName": "夜鸟类",
        "parentId": "strisores_phaethoquornithes",
        "summary": "包括油鸱、林鸱、蟆口鸱、裸鼻鸱、夜鹰、雨燕和蜂鸟等。",
        "traits": ["包含夜行、暮行、高速飞行和花蜜取食等多种特化", "蜂鸟的悬停飞行与雨燕的高速飞行都落在这个大分支内"],
    },
    {
        "id": "aequornithes",
        "rank": "clade",
        "scientificName": "Aequornithes",
        "englishName": "Core waterbirds",
        "chineseName": "核心水鸟类",
        "parentId": "phaethoquornithes",
        "summary": "鹭形类内部的一组与水域、海洋或湿地生态高度相关的鸟类分支。",
        "traits": ["潜水、海洋飞行、涉水取食等水域适应反复出现", "包含企鹅、鹱、潜鸟、鲣鸟、鹈鹕等"],
    },
    {
        "id": "mirandornithes",
        "rank": "clade",
        "scientificName": "Mirandornithes",
        "englishName": "Flamingos and grebes",
        "chineseName": "奇迹鸟类",
        "parentId": "neoaves",
        "summary": "红鹳和䴙䴘的共同分支，是“外形直觉容易误导”的经典例子之一。",
        "traits": ["红鹳涉水滤食，䴙䴘潜水捕鱼，外形差异大但亲缘接近"],
    },
    {
        "id": "columbaves",
        "rank": "clade",
        "scientificName": "Columbaves",
        "englishName": "Pigeons, sandgrouse, and allies",
        "chineseName": "鸽鸨类",
        "parentId": "neoaves",
        "summary": "包含鸽形目、沙鸡目和拟鹑目等，内部关系在不同研究中会有细节差异。",
        "traits": ["包含种子食性、地栖和树栖等多种生活方式"],
    },
    {
        "id": "otidimorphae",
        "rank": "clade",
        "scientificName": "Otidimorphae",
        "englishName": "Bustards, turacos, and cuckoos",
        "chineseName": "鸨鹃类",
        "parentId": "neoaves",
        "summary": "将鸨、蕉鹃和鹃类放在一个较高层级的演化支中。",
        "traits": ["外形差异很大，适合展示系统发育与直觉形态分类的差别"],
    },
    {
        "id": "elementaves",
        "rank": "clade",
        "scientificName": "Elementaves",
        "englishName": "Elementaves",
        "chineseName": "元素鸟类",
        "parentId": "neoaves",
        "summary": "近年在万种鸟类基因组计划中被最新描述的一大类群。下属分支包括 Gruae，以及夜鸟类与鹭形类所在的姐妹支。",
        "traits": ["把麝雉目、鹤形总目、夜鸟类、日鳽目、鹲形目与核心水鸟类放入同一大框架中"],
    },
    {
        "id": "gruae",
        "rank": "clade",
        "scientificName": "Gruae",
        "englishName": "Hoatzin, cranes, and shorebirds",
        "chineseName": "Gruae",
        "parentId": "elementaves",
        "summary": "元素鸟类内部的一支，包含麝雉目 Opisthocomiformes 与鹤形总目 Cursorimorphae。",
        "traits": ["在展示层中把麝雉与鹤形目、鸻形目所在分支并列连接"],
    },
    {
        "id": "cursorimorphae",
        "rank": "clade",
        "scientificName": "Cursorimorphae",
        "englishName": "Cranes and shorebirds",
        "chineseName": "鹤形总目",
        "parentId": "gruae",
        "summary": "由鹤形目与鸻形目共同构成的演化支。",
        "traits": ["湿地、岸滩、草地和水边生态多样", "包含许多长腿涉禽、岸鸟和相关类群"],
    },
    {
        "id": "strisores_phaethoquornithes",
        "rank": "clade",
        "scientificName": "Strisores + Phaethoquornithes",
        "englishName": "Nightbirds and Phaethoquornithes",
        "chineseName": "夜鸟类-鹭形类支",
        "parentId": "elementaves",
        "summary": "元素鸟类内部由夜鸟类与鹭形类组成的一支，目前可先按这两个组成类群来理解。",
        "traits": ["连接夜鸟类与日鳽、鹲类及核心水鸟类所在分支"],
    },
    {
        "id": "phaethoquornithes",
        "rank": "clade",
        "scientificName": "Phaethoquornithes",
        "englishName": "Sunbittern, tropicbirds, and core waterbirds",
        "chineseName": "鹭形类",
        "parentId": "strisores_phaethoquornithes",
        "summary": "由日鳽总目 Phaethontimorphae 与核心水鸟类 Aequornithes 共同构成的演化支。",
        "traits": ["连接日鳽、鹲类与更大范围的水鸟谱系"],
    },
    {
        "id": "phaethontimorphae",
        "rank": "clade",
        "scientificName": "Phaethontimorphae",
        "englishName": "Sunbittern and tropicbirds",
        "chineseName": "日鳽总目",
        "parentId": "phaethoquornithes",
        "summary": "包含日鳽目 Eurypygiformes 与鹲形目 Phaethontiformes 的演化支。",
        "traits": ["日鳽与鹲类外形和生态差异明显，但在这里作为相邻谱系展示"],
    },
    {
        "id": "afroaves",
        "rank": "clade",
        "scientificName": "Afroaves",
        "englishName": "Afroaves",
        "chineseName": "非洲鸟类",
        "parentId": "telluraves",
        "summary": "陆鸟类中的一支，包含鹰形目、鸮形目、犀鸟、佛法僧、啄木鸟等多组。",
        "traits": ["视觉捕猎、树栖、洞巢和攀缘等生态形态多样"],
    },
    {
        "id": "australaves",
        "rank": "clade",
        "scientificName": "Australaves",
        "englishName": "Australaves",
        "chineseName": "澳洲鸟类",
        "parentId": "telluraves",
        "summary": "陆鸟类中的一支，包含隼形目、鹦形目、雀形目和叫鹤目等。",
        "traits": ["隼与鹦鹉、雀形鸟同在这一大支内，是猛禽趋同学习的关键节点"],
    },
    {
        "id": "telluraves",
        "rank": "clade",
        "scientificName": "Telluraves",
        "englishName": "Landbirds",
        "chineseName": "陆鸟类",
        "parentId": "neoaves",
        "summary": "大型陆生鸟类演化支，包含猛禽、鹦鹉、雀形鸟、啄木鸟、犀鸟等。",
        "traits": ["许多树栖、捕食和鸣唱生态位在此分支内辐射", "隼形目与鹰形目相似但不相邻"],
    },
]


FOCUS_SPECIES = {
    "Common Nighthawk",
    "Eastern Whip-poor-will",
    "Oilbird",
    "Common Potoo",
    "Tawny Frogmouth",
    "Australian Owlet-nightjar",
    "Ruby-throated Hummingbird",
    "Sword-billed Hummingbird",
    "Chimney Swift",
    "Gray-rumped Treeswift",
    "Barn Swallow",
    "Purple Sunbird",
    "Peregrine Falcon",
    "Crested Caracara",
    "Bald Eagle",
    "Osprey",
    "Secretarybird",
    "Western Barn Owl",
    "Great Horned Owl",
    "Emperor Penguin",
    "King Penguin",
    "Atlantic Puffin",
    "Common Murre",
    "Chicken",
    "Mallard",
    "Common Ostrich",
    "Emu",
    "Rifleman",
    "Superb Lyrebird",
    "American Crow",
    "Zebra Finch",
    "Kea",
}


NODE_NOTES = {
    "Trochilidae": {
        "chineseName": "蜂鸟科",
        "summary": "现代蜂鸟多样性集中在美洲，具有极端飞行控制和花蜜取食适应。",
        "traits": ["鸟类中最典型的持续悬停飞行", "长喙和舌部结构与花蜜取食高度相关", "常与太阳鸟产生趋同外观"],
    },
    "Apodidae": {
        "chineseName": "雨燕科",
        "summary": "高度适应空中生活的类群，与燕子相似但亲缘不近。",
        "traits": ["长时间空中飞行", "镰刀状翅形常见", "与蜂鸟同属雨燕目"],
    },
    "Caprimulgidae": {
        "chineseName": "夜鹰科",
        "summary": "多为暮行或夜行，张口捕食飞虫，是夜鸟类的重要成员。",
        "traits": ["隐蔽斑驳羽色", "夜间或黄昏活动", "宽口裂适合空中捕虫"],
    },
    "Steatornithidae": {
        "chineseName": "油鸱科",
        "summary": "油鸱是独特的夜行果食鸟类，使用回声定位，是夜鸟类中的特殊分支。",
        "traits": ["夜行果食", "洞穴栖息", "使用可听声回声定位"],
    },
    "Nyctibiidae": {
        "chineseName": "林鸱科",
        "summary": "林鸱以拟态停栖著称，外形和行为高度适应隐蔽生活。",
        "traits": ["木桩状拟态", "夜行捕虫", "大眼和宽口裂"],
    },
    "Podargidae": {
        "chineseName": "蟆口鸱科",
        "summary": "蟆口鸱与夜鹰相似，但在现代分类中通常单列为不同目。",
        "traits": ["宽阔蟆状嘴", "夜行或暮行", "拟态与伏击取食"],
    },
    "Aegothelidae": {
        "chineseName": "裸鼻鸱科",
        "summary": "裸鼻鸱是夜鸟类中较小而隐蔽的分支，多分布于澳新区域。",
        "traits": ["夜行", "树洞利用", "联系夜鹰与雨燕蜂鸟方向的学习节点"],
    },
    "Falconidae": {
        "chineseName": "隼科",
        "summary": "隼、卡拉鹰等属于隼形目；它们像鹰，但与鹦鹉和雀形鸟所在大支更接近。",
        "traits": ["高速追击", "钩喙与强视觉捕猎", "与鹰形目相似主要来自趋同演化"],
    },
    "Accipitridae": {
        "chineseName": "鹰科",
        "summary": "鹰、雕、鹞、鸢等典型日行猛禽所在科。",
        "traits": ["钩喙利爪", "强视觉捕猎", "许多种有明显眼眶上脊形成猛禽脸"],
    },
    "Pandionidae": {
        "chineseName": "鹗科",
        "summary": "鹗是高度特化的捕鱼猛禽，在 Clements 中作为鹰形目内独立科。",
        "traits": ["鱼食性", "全球广布", "脚部结构适合抓鱼"],
    },
    "Sagittariidae": {
        "chineseName": "蛇鹫科",
        "summary": "蛇鹫是非洲草原的长腿猛禽，形态非常独特。",
        "traits": ["长腿步行捕食", "常捕食蛇类和小动物", "鹰形目中的独特谱系"],
    },
    "Hirundinidae": {
        "chineseName": "燕科",
        "summary": "燕子属于雀形目，与雨燕外形和生态相似但亲缘不近。",
        "traits": ["空中捕虫", "叉尾和长翼常见", "与雨燕是趋同案例"],
    },
    "Nectariniidae": {
        "chineseName": "太阳鸟科",
        "summary": "太阳鸟是旧大陆花蜜取食雀形鸟，常与蜂鸟相似但亲缘较远。",
        "traits": ["花蜜取食", "金属色羽毛常见", "通常不能像蜂鸟那样持续悬停"],
    },
    "Spheniscidae": {
        "chineseName": "企鹅科",
        "summary": "企鹅是南半球海鸟，翅膀特化为水下推进结构。",
        "traits": ["不会飞", "水下推进", "与北半球海雀外形相似但亲缘不同"],
    },
    "Alcidae": {
        "chineseName": "海雀科",
        "summary": "海雀、海鹦和崖海鸦等属于鸻形目，是北半球海鸟。",
        "traits": ["潜水捕食", "黑白海鸟外观", "与企鹅形成经典趋同对照"],
    },
    "Passeriformes": {
        "summary": "雀形目是现代鸟类最大目，包含鸣禽和亚鸣禽等重要分支。",
        "traits": ["鸣唱、树栖和小型化多样性极高", "oscines 起源与澳大拉西亚相关，是后续演化历史模块的重点"],
    },
}


COMPARISONS = [
    {
        "title": "隼形目 vs 鹰形目",
        "leftId": "order_falconiformes",
        "rightId": "order_accipitriformes",
        "contextIds": ["telluraves", "australaves", "afroaves", "order_falconiformes", "order_accipitriformes"],
        "summary": "两者都有钩喙、利爪和强视觉捕猎，但隼形目在陆鸟类中更接近鹦鹉与雀形鸟方向；鹰形目属于另一支日行猛禽谱系。",
        "shared": ["日行捕猎生态位", "钩喙", "高速视觉追踪"],
        "difference": "相似外形主要来自趋同演化，不代表最近亲缘。",
    },
    {
        "title": "蜂鸟 vs 太阳鸟",
        "leftId": "family_trochilidae",
        "rightId": "family_nectariniidae",
        "contextIds": ["strisores", "order_apodiformes", "family_trochilidae", "order_passeriformes", "family_nectariniidae"],
        "summary": "两者都与花蜜取食、长喙和艳丽羽色相关，但蜂鸟在夜鸟类/雨燕目内，太阳鸟是雀形目。",
        "shared": ["花蜜取食", "长喙", "鲜艳羽色"],
        "difference": "蜂鸟拥有极端悬停飞行能力；太阳鸟多以停栖取食为主。",
    },
    {
        "title": "雨燕 vs 燕子",
        "leftId": "family_apodidae",
        "rightId": "family_hirundinidae",
        "contextIds": ["strisores", "order_apodiformes", "family_apodidae", "order_passeriformes", "family_hirundinidae"],
        "summary": "雨燕和燕子都在空中捕虫，外形近似，但雨燕与蜂鸟同属雨燕目，燕子属于雀形目。",
        "shared": ["空中捕虫", "长翼", "高速飞行"],
        "difference": "外形相似来自相同飞行生态位，而不是近缘关系。",
    },
    {
        "title": "企鹅 vs 海雀",
        "leftId": "family_spheniscidae",
        "rightId": "family_alcidae",
        "contextIds": ["elementaves", "phaethoquornithes", "aequornithes", "order_sphenisciformes", "family_spheniscidae", "cursorimorphae", "order_charadriiformes", "family_alcidae"],
        "summary": "企鹅和海雀都是黑白海鸟并擅长潜水，但企鹅属于企鹅目，海雀属于鸻形目。",
        "shared": ["海洋生活", "潜水捕食", "黑白反差羽色"],
        "difference": "企鹅不会飞且南半球为主；多数海雀仍会飞，北半球为主。",
    },
    {
        "title": "鸮形目 vs 日行猛禽",
        "leftId": "order_strigiformes",
        "rightId": "order_accipitriformes",
        "contextIds": ["telluraves", "afroaves", "order_strigiformes", "order_accipitriformes"],
        "summary": "猫头鹰和鹰雕都占据猛禽生态位，但猫头鹰是夜行捕猎谱系，鹰雕是日行猛禽谱系。",
        "shared": ["钩喙", "利爪", "捕食脊椎动物或大型猎物"],
        "difference": "鸮形目具有夜视、静音飞行和面盘等夜行适应。",
    },
]


def slug(value):
    return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")


def parse_family_label(label):
    m = re.match(r"^(.+?) \((.+)\)$", label)
    if m:
        return m.group(1), m.group(2)
    return label, ""

def sort_value(row):
    try:
        return float(row.get("Sequence") or row.get("sort v2025") or 999999)
    except (TypeError, ValueError):
        return 999999


def load_avilist_rows():
    wb = load_workbook(AVILIST_PATH, read_only=True, data_only=True)
    ws = wb["AviList v2025b extended"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for values in ws.iter_rows(min_row=3, values_only=True):
        row = dict(zip(headers, values))
        if row.get("Taxon_rank"):
            rows.append(row)
    return rows


def load_genus_chinese_overrides():
    overrides = {}
    processed_dir = ROOT / "data" / "processed"
    paths = sorted({
        *processed_dir.glob("*genus*filled.csv"),
        *processed_dir.glob("*genera*filled.csv"),
    })
    final_reviewed_paths = sorted(processed_dir.glob("*final-reviewed*.csv"))
    final_reviewed_set = set(final_reviewed_paths)
    for path in paths:
        if path in final_reviewed_set:
            continue
        with path.open(newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                genus = (row.get("Genus") or "").strip()
                chinese = (
                    row.get("Official Chinese genus name")
                    or row.get("Suggested Chinese genus name")
                    or ""
                ).strip()
                if genus and chinese:
                    overrides[genus] = chinese
    for path in final_reviewed_paths:
        with path.open(newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                genus = (row.get("Genus") or "").strip()
                chinese = (
                    row.get("Official Chinese genus name")
                    or row.get("Suggested Chinese genus name")
                    or ""
                ).strip()
                if genus:
                    if chinese:
                        overrides[genus] = chinese
                    else:
                        overrides.pop(genus, None)
    return overrides


def load_china_checklist_rows():
    book = xlrd.open_workbook(CHINA_CHECKLIST_PATH)
    sheet = book.sheet_by_name("Checklist_12.0")
    rows = []
    for row_index in range(8, sheet.nrows):
        values = [sheet.cell_value(row_index, col) for col in range(sheet.ncols)]
        scientific = str(values[1]).strip()
        chinese = str(values[2]).strip()
        english = str(values[3]).strip()
        if scientific and chinese and english and re.search(r"[A-Z][a-z]+\s+[a-z]", scientific):
            rows.append(
                {
                    "sourceRow": row_index + 1,
                    "scientificName": scientific,
                    "chineseName": chinese,
                    "englishName": english,
                    "iucn": str(values[4]).strip(),
                    "protection": str(values[5]).strip(),
                    "notes": str(values[6]).strip(),
                }
            )
    return rows


def load_china_family_names():
    book = xlrd.open_workbook(CHINA_CHECKLIST_PATH)
    sheet = book.sheet_by_name("Checklist_12.0")
    family_names = {}
    report_rows = []
    for row_index in range(8, sheet.nrows):
        values = [str(sheet.cell_value(row_index, col)).strip() for col in range(sheet.ncols)]
        order = values[1]
        chinese = values[2]
        raw_family = values[3]
        match = re.match(r"^([A-Z][a-z]+idae)\b", raw_family)
        if order.isupper() and chinese and match:
            family = match.group(1)
            family_names[family] = {
                "chineseName": chinese,
                "source": "CBR Checklist of Birds of China v12.0 (2024)",
                "sourceRow": row_index + 1,
                "orderOriginal": order,
                "raw": raw_family,
            }
            report_rows.append(
                {
                    "CBR row": row_index + 1,
                    "CBR order": order,
                    "Family": family,
                    "Chinese family name": chinese,
                    "Raw family cell": raw_family,
                }
            )
    return family_names, report_rows


def build_china_matches(china_rows, avilist_species_rows):
    by_scientific = {str(row["Scientific_name"]): row for row in avilist_species_rows}
    by_english = defaultdict(dict)
    for row in avilist_species_rows:
        for field in ["English_name_AviList", "English_name_Clements_v2025", "English_name_BirdLife_v10"]:
            name = row.get(field)
            if name:
                by_english[str(name).strip().lower()][str(row["Scientific_name"])] = row

    matches = {}
    report_rows = []
    for china in china_rows:
        match = by_scientific.get(china["scientificName"])
        method = "scientificName"
        status = "matched_exact"
        candidates = []
        if not match:
            candidates = list(by_english.get(china["englishName"].lower(), {}).values())
            if len(candidates) == 1:
                match = candidates[0]
                method = "englishName"
                status = "matched_name_conflict_avilist_taxonomy_used"
            else:
                method = "englishName"
                status = "unmatched" if not candidates else "ambiguous_english_name"

        avilist_name = str(match["Scientific_name"]) if match else ""
        if match:
            payload = {
                "source": "CBR Checklist of Birds of China v12.0 (2024)",
                "sourceRow": china["sourceRow"],
                "matchMethod": method,
                "matchStatus": status,
                "scientificNameOriginal": china["scientificName"],
                "chineseName": china["chineseName"],
                "englishName": china["englishName"],
                "iucn": china["iucn"],
                "protection": china["protection"],
                "notes": china["notes"],
            }
            matches[avilist_name] = payload

        report_rows.append(
            {
                "status": status,
                "matchMethod": method,
                "CBR row": china["sourceRow"],
                "CBR scientific name": china["scientificName"],
                "CBR Chinese name": china["chineseName"],
                "CBR English name": china["englishName"],
                "AviList scientific name used": avilist_name,
                "AviList English name": match.get("English_name_AviList", "") if match else "",
                "candidate scientific names": "; ".join(str(row["Scientific_name"]) for row in candidates[:10]),
                "CBR notes": china["notes"],
            }
        )
    return matches, report_rows


def write_china_match_report(report_rows):
    out = ROOT / "data" / "processed" / "china-checklist-avilist-match-report.csv"
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=list(report_rows[0].keys()))
        writer.writeheader()
        writer.writerows(report_rows)


def write_china_family_name_report(report_rows, avilist_family_names):
    out = ROOT / "data" / "processed" / "china-family-name-avilist-match-report.csv"
    out.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["status", "CBR row", "CBR order", "Family", "Chinese family name", "Raw family cell"]
    enriched = []
    for row in report_rows:
        enriched.append(
            {
                "status": "matched" if row["Family"] in avilist_family_names else "unmatched",
                **row,
            }
        )
    with out.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(enriched)


def common_suffix(values):
    if not values:
        return ""
    suffix = values[0]
    for value in values[1:]:
        i = 0
        max_len = min(len(suffix), len(value))
        while i < max_len and suffix[-1 - i] == value[-1 - i]:
            i += 1
        suffix = suffix[len(suffix) - i :] if i else ""
        if not suffix:
            break
    return suffix


def derive_genus_names_from_china_checklist(china_matches):
    by_genus = defaultdict(list)
    for avilist_scientific, china in china_matches.items():
        genus = avilist_scientific.split()[0]
        chinese = (china.get("chineseName") or "").strip()
        if chinese:
            by_genus[genus].append(
                {
                    "scientificName": avilist_scientific,
                    "chineseName": chinese,
                    "sourceRow": china.get("sourceRow", ""),
                }
            )

    derived = {}
    report_rows = []
    for genus, species in sorted(by_genus.items()):
        names = [row["chineseName"] for row in species]
        suffix = common_suffix(names)
        accepted = len(species) >= 2 and len(suffix) >= 2
        suggested = f"{suffix}属" if accepted else ""
        if accepted:
            derived[genus] = {
                "chineseName": suggested,
                "source": "Derived from shared suffix in CBR Checklist v12.0 Chinese species names",
                "method": "common Chinese species-name suffix; no English/Latin semantic translation",
                "speciesCount": len(species),
                "suffix": suffix,
            }
        report_rows.append(
            {
                "status": "accepted" if accepted else "not_used",
                "Genus": genus,
                "Suggested Chinese genus name": suggested,
                "Shared suffix": suffix,
                "CBR species count": len(species),
                "Example species": "; ".join(
                    f"{row['chineseName']} ({row['scientificName']})" for row in species[:8]
                ),
                "Method": "common Chinese species-name suffix; minimum 2 CBR species and suffix length >= 2",
            }
        )
    return derived, report_rows


def write_cbr_derived_genus_report(report_rows):
    out = ROOT / "data" / "processed" / "cbr-derived-genus-chinese-names.csv"
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", newline="", encoding="utf-8-sig") as f:
        fieldnames = [
            "status",
            "Genus",
            "Suggested Chinese genus name",
            "Shared suffix",
            "CBR species count",
            "Example species",
            "Method",
        ]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(report_rows)


def write_cbr_missing_official_genus_report(china_matches, species_rows, genus_zh, derived_report_rows, china_family_names):
    out = ROOT / "data" / "processed" / "cbr-genera-needing-official-chinese-names.csv"
    out.parent.mkdir(parents=True, exist_ok=True)
    species_by_name = {str(row["Scientific_name"]): row for row in species_rows}
    derived_by_genus = {row["Genus"]: row for row in derived_report_rows}
    by_genus = defaultdict(list)
    for scientific, china in china_matches.items():
        row = species_by_name.get(scientific)
        if not row:
            continue
        genus = scientific.split()[0]
        if genus_zh.get(genus):
            continue
        by_genus[genus].append((scientific, china, row))

    report_rows = []
    for genus, entries in sorted(by_genus.items(), key=lambda item: (-len(item[1]), item[0])):
        first_row = entries[0][2]
        family = first_row["Family"]
        derived = derived_by_genus.get(genus, {})
        report_rows.append(
            {
                "Order": first_row["Order"],
                "Order Chinese": ORDER_ZH.get(first_row["Order"], ""),
                "Family": family,
                "Family Chinese": (china_family_names.get(family) or {}).get("chineseName", ""),
                "Genus": genus,
                "Official Chinese genus name": "",
                "CBR suffix candidate - NOT USED": derived.get("Suggested Chinese genus name", ""),
                "Candidate basis": "shared Chinese species-name suffix only; not used in app"
                if derived.get("Suggested Chinese genus name")
                else "",
                "CBR checklist species count": len(entries),
                "Example CBR species": "; ".join(
                    f"{china.get('chineseName', '')} ({scientific})" for scientific, china, _ in entries[:10]
                ),
            }
        )

    fieldnames = [
        "Order",
        "Order Chinese",
        "Family",
        "Family Chinese",
        "Genus",
        "Official Chinese genus name",
        "CBR suffix candidate - NOT USED",
        "Candidate basis",
        "CBR checklist species count",
        "Example CBR species",
    ]
    with out.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(report_rows)


def load_chinese_map():
    text = CHINESE_PATH.read_text(encoding="utf-8", errors="replace")
    simple = {}
    for english, full, chinese in re.findall(
        r'"([^"\\]*(?:\\.[^"\\]*)*)"\s*:\s*"([^"\\]*(?:\\.[^"\\]*)*)\(([^()]+)\)"',
        text,
    ):
        simple[english] = chinese
    structured = {}
    structured_by_code = {}
    pattern = re.compile(
        r'"([^"\\]*(?:\\.[^"\\]*)*)"\s*:\s*\{\s*"pinyin"\s*:\s*"([^"]*)"\s*,\s*"initials"\s*:\s*"([^"]*)"\s*,\s*"code"\s*:\s*"([^"]*)"\s*,\s*"name"\s*:\s*"([^"]*)"\s*,\s*"latin"\s*:\s*"([^"]*)"',
        re.S,
    )
    for key, pinyin, initials, code, name, latin in pattern.findall(text):
        structured[name] = {
            "chineseName": simple.get(name, ""),
            "pinyin": pinyin,
            "initials": initials,
            "code": code,
            "scientificName": latin,
        }
        if code:
            structured_by_code[code] = structured[name]
    return simple, structured, structured_by_code


def main():
    APP_DIR.mkdir(parents=True, exist_ok=True)
    simple_zh, structured_zh, structured_by_code = load_chinese_map()
    genus_zh = load_genus_chinese_overrides()
    rows = load_avilist_rows()

    order_rows = [r for r in rows if r["Taxon_rank"] == "order"]
    family_rows = [r for r in rows if r["Taxon_rank"] == "family"]
    genus_rows = [r for r in rows if r["Taxon_rank"] == "genus"]
    species_rows = [r for r in rows if r["Taxon_rank"] == "species"]
    china_rows = load_china_checklist_rows()
    china_matches, china_match_report = build_china_matches(china_rows, species_rows)
    write_china_match_report(china_match_report)
    cbr_genus_zh, cbr_genus_report = derive_genus_names_from_china_checklist(china_matches)
    write_cbr_derived_genus_report(cbr_genus_report)
    china_family_names, china_family_report = load_china_family_names()
    avilist_family_names = {str(row["Scientific_name"]) for row in family_rows}
    write_china_family_name_report(china_family_report, avilist_family_names)
    write_cbr_missing_official_genus_report(
        china_matches, species_rows, genus_zh, cbr_genus_report, china_family_names
    )
    species_by_family = defaultdict(list)
    for row in species_rows:
        species_by_family[row["Family"]].append(row)
    for family in species_by_family:
        species_by_family[family].sort(key=sort_value)

    order_sort = {}
    for row in order_rows:
        order = row["Scientific_name"]
        if order:
            order_sort[order] = min(order_sort.get(order, 999999), sort_value(row))

    nodes = []
    for clade in CLADES:
        nodes.append(
            {
                **clade,
                "sortOrder": CLADE_SORT.get(clade["id"], 999),
                "sortBasis": "Manual prototype clade order",
                "source": "Manual prototype phylogeny layer",
            }
        )

    order_nodes = set()
    for row in sorted(order_rows, key=sort_value):
        order = row["Scientific_name"]
        node_id = f"order_{slug(order)}"
        parent = ORDER_PARENT.get(order, "neoaves_other")
        order_nodes.add(order)
        note = NODE_NOTES.get(order, {})
        nodes.append(
            {
                "id": node_id,
                "rank": "order",
                "scientificName": order,
                "englishName": order,
                "chineseName": ORDER_ZH.get(order, ""),
                "parentId": parent,
                "summary": note.get("summary", ""),
                "traits": note.get("traits", []),
                "sortOrder": order_sort.get(order, 999999),
                "sortBasis": "AviList Sequence",
                "source": "AviList v2025b; Chinese order names manually seeded",
            }
        )

    family_id_by_label = {}
    family_id_by_scientific = {}
    for row in sorted(family_rows, key=sort_value):
        scientific = row["Scientific_name"]
        english = row.get("Family_English_name") or ""
        node_id = f"family_{slug(scientific)}"
        family_id_by_label[row["Family"]] = node_id
        family_id_by_scientific[scientific] = node_id
        note = NODE_NOTES.get(scientific, {})
        china_family = china_family_names.get(scientific)
        nodes.append(
            {
                "id": node_id,
                "rank": "family",
                "scientificName": scientific,
                "englishName": english or row["English name"],
                "chineseName": china_family.get("chineseName") if china_family else note.get("chineseName", ""),
                "parentId": f"order_{slug(row['Order'])}",
                "summary": note.get("summary", ""),
                "traits": note.get("traits", []),
                "speciesCountClements": len(species_by_family[row["Family"]]),
                "chinaFamilyName": china_family,
                "sortOrder": sort_value(row),
                "sortBasis": "AviList Sequence family row",
                "source": "AviList v2025b family record; Chinese family name from CBR v12.0 when matched",
            }
        )

    genus_id_by_scientific = {}
    for row in sorted(genus_rows, key=sort_value):
        scientific = row["Scientific_name"]
        node_id = f"genus_{slug(scientific)}"
        genus_id_by_scientific[scientific] = node_id
        manual_genus_name = genus_zh.get(scientific, "")
        derived_genus = cbr_genus_zh.get(scientific)
        nodes.append(
            {
                "id": node_id,
                "rank": "genus",
                "scientificName": scientific,
                "englishName": "",
                "chineseName": manual_genus_name,
                "parentId": family_id_by_scientific.get(row["Family"]),
                "summary": "",
                "traits": [],
                "chinaGenusNameCandidate": derived_genus,
                "sortOrder": sort_value(row),
                "sortBasis": "AviList Sequence genus row",
                "source": "AviList v2025b genus record; Chinese genus name from local override when matched",
            }
        )

    china_terms = re.compile(
        r"China|Chinese|Taiwan|Hong Kong|Macau|Tibet|Xinjiang|Manchuria|Hainan|Yunnan|Sichuan|Qinghai|Inner Mongolia",
        re.I,
    )

    for row in sorted(species_rows, key=sort_value):
        avilist_name = row.get("English_name_AviList") or ""
        clements_name = row.get("English_name_Clements_v2025") or ""
        species_code = row.get("Species_code_Cornell_Lab") or ""
        zh = structured_by_code.get(species_code) or structured_zh.get(clements_name) or structured_zh.get(avilist_name) or {}
        chinese = zh.get("chineseName") or simple_zh.get(clements_name) or simple_zh.get(avilist_name) or ""
        genus_name = str(row["Scientific_name"]).split()[0]
        range_text = row.get("Range") or ""
        region_tags = ["China range keyword match"] if range_text and china_terms.search(str(range_text)) else []
        china_checklist = china_matches.get(str(row["Scientific_name"]))
        if china_checklist:
            if "China checklist v12.0" not in region_tags:
                region_tags.insert(0, "China checklist v12.0")
            chinese = china_checklist["chineseName"] or chinese
        nodes.append(
            {
                "id": f"species_{slug(row['Scientific_name'])}",
                "rank": "species",
                "scientificName": row["Scientific_name"],
                "englishName": avilist_name or clements_name,
                "englishNameClements": clements_name,
                "englishNameBirdLife": row.get("English_name_BirdLife_v10") or "",
                "chineseName": chinese,
                "parentId": genus_id_by_scientific.get(genus_name) or family_id_by_scientific.get(row["Family"]),
                "range": range_text,
                "ebirdCode": species_code,
                "avibaseId": row.get("AvibaseID") or "",
                "iucn": row.get("IUCN_Red_List_Category") or "",
                "birdLifeUrl": row.get("BirdLife_DataZone_URL") or "",
                "birdsOfTheWorldUrl": row.get("Birds_of_the_World_URL") or "",
                "decisionSummary": row.get("Decision_summary") or "",
                "regionTags": region_tags,
                "chinaChecklist": china_checklist,
                "pinyin": zh.get("pinyin", ""),
                "initials": zh.get("initials", ""),
                "summary": "",
                "traits": [],
                "sortOrder": sort_value(row),
                "sortBasis": "AviList Sequence species row",
                "source": "AviList v2025b species record; Chinese name from birdMapV2.js when matched",
            }
        )

    children = defaultdict(list)
    node_by_id = {n["id"]: n for n in nodes}
    for node in nodes:
        pid = node.get("parentId")
        if pid:
            children[pid].append(node["id"])

    rank_order = {"class": 0, "subclass": 1, "clade": 2, "order": 3, "family": 4, "species": 5}
    for node in nodes:
        node["childrenIds"] = sorted(
            children.get(node["id"], []),
            key=lambda cid: (
                rank_order.get(node_by_id[cid]["rank"], 9),
                node_by_id[cid].get("sortOrder", 999999),
                node_by_id[cid]["scientificName"],
            ),
        )

    data = {
        "meta": {
            "title": "Bird Phylogeny Explorer",
            "subtitle": "鸟类谱系速查器",
            "prototypeDate": "2026-07-24",
            "clementsFile": CLEMENTS_PATH.name,
            "avilistFile": AVILIST_PATH.name,
            "avilistRows": len(rows),
            "includedNodes": len(nodes),
            "includedOrders": len([n for n in nodes if n["rank"] == "order"]),
            "includedFamilies": len([n for n in nodes if n["rank"] == "family"]),
            "includedFamiliesWithChinese": len([n for n in nodes if n["rank"] == "family" and n.get("chineseName")]),
            "chinaFamilyRows": len(china_family_report),
            "chinaFamilyRowsMatchedAviList": len([r for r in china_family_report if r["Family"] in avilist_family_names]),
            "includedGenera": len([n for n in nodes if n["rank"] == "genus"]),
            "includedGeneraWithChinese": len([n for n in nodes if n["rank"] == "genus" and n.get("chineseName")]),
            "includedGeneraWithManualChinese": len(genus_zh),
            "includedGeneraWithCbrDerivedCandidates": len(cbr_genus_zh),
            "includedSpecies": len([n for n in nodes if n["rank"] == "species"]),
            "chinaRangeKeywordSpecies": len([n for n in nodes if "China range keyword match" in n.get("regionTags", [])]),
            "chinaChecklistRows": len(china_rows),
            "chinaChecklistMatchedSpecies": len(china_matches),
            "chinaChecklistUnmatchedRows": len([r for r in china_match_report if r["status"] in {"unmatched", "ambiguous_english_name"}]),
            "sources": [
                "AviList Core Team. 2026. AviList: The Global Avian Checklist, v2025b. https://doi.org/10.2173/avilist.v2025b",
                "中国观鸟年报-中国鸟类名录 12.0 (2024), used for China checklist tags and Chinese species names; AviList taxonomy prevails when concepts differ.",
                "birdMapV2.js by wzy0421, used as prototype Chinese species-name mapping.",
                "Clements/eBird v2025 names retained where AviList provides perfect-match Clements names.",
                "Prum et al. 2015, Nature, doi:10.1038/nature15697.",
                "Stiller et al. 2024, Nature, doi:10.1038/s41586-024-07323-1.",
            ],
            "warning": "High-level clade placement is a prototype learning layer and should be reviewed before publication.",
            "sortPolicy": "Sibling order uses AviList Sequence for imported orders, families, genera, and species; manual clades use prototype phylogenetic display order.",
        },
        "nodes": nodes,
        "comparisons": COMPARISONS,
        "quickStarts": [
            "美洲夜鹰",
            "红喉北蜂鸟",
            "游隼",
            "白头海雕",
            "家燕",
            "紫色花蜜鸟",
            "帝企鹅",
            "北极海鹦",
            "Strisores",
            "Galloanserae",
            "Passeriformes",
        ],
    }

    out = APP_DIR / "bird-data.js"
    out.write_text("window.BIRD_DATA = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n", encoding="utf-8")
    print(json.dumps(data["meta"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
